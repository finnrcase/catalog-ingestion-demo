import io

import pandas as pd
import pytest

from src.programa_export import (
    _clean_notes,
    _extract_material_from_notes,
)


# ── Shared test fixtures ──────────────────────────────────────────────────────

def _scotsman_row() -> dict:
    """Acceptance test fixture — Scotsman icemaker."""
    return {
        "Include": True,
        "Product Category": "Appliances",
        "Product Name": "Scotsman Icemaker Built-In Pump",
        "Brand": "Scotsman",
        "Model/SKU": "SCN60PA1SU",
        "Dimensions": "14.875 in W x 22 in D x 33.375 in H",
        "Finish / Color": "Stainless Steel",
        "Quantity": 1,
        "Price": "",
        "Supplier": "Scotsman",
        "Product URL": "https://scotsman-ice.com/product/scn60pa1su",
        "Image URL": "https://scotsman-ice.com/images/scn60pa1su.jpg",
        "Notes": "Verify delivery date. [Materials: Stainless Steel] [Enrichment: matched]",
        "Room": "Kitchen",
    }


def _make_rows(overrides_list: list[dict]) -> list[dict]:
    """Build a list of minimal valid rows, applying per-row overrides."""
    base = {
        "Include": True,
        "Product Name": "Test Product",
        "Product Category": "Lighting",
        "Dimensions": "12 in W x 10 in H x 8 in D",
        "Product URL": "https://example.com/product",
        "Image URL": "https://example.com/image.jpg",
    }
    return [{**base, **o} for o in overrides_list]


# ── _extract_material_from_notes ──────────────────────────────────────────────

def test_extract_material_from_tag():
    assert _extract_material_from_notes("[Materials: Stainless Steel]") == "Stainless Steel"

def test_extract_material_case_insensitive():
    assert _extract_material_from_notes("[materials: oak veneer]") == "oak veneer"

def test_extract_material_trims_whitespace():
    assert _extract_material_from_notes("[Materials:  solid oak  ]") == "solid oak"

def test_extract_material_with_surrounding_text():
    result = _extract_material_from_notes("Some note. [Materials: Brass] More text.")
    assert result == "Brass"

def test_extract_material_missing_tag_returns_empty():
    assert _extract_material_from_notes("Just a regular note.") == ""

def test_extract_material_empty_string():
    assert _extract_material_from_notes("") == ""


# ── _clean_notes ──────────────────────────────────────────────────────────────

def test_clean_notes_strips_materials_tag():
    result = _clean_notes("Good note. [Materials: Steel]")
    assert "[Materials:" not in result
    assert "Good note." in result

def test_clean_notes_strips_enrichment_tag():
    result = _clean_notes("[Enrichment: no confident source found] Some note.")
    assert "[Enrichment:" not in result
    assert "Some note." in result

def test_clean_notes_strips_partial_dimension_tag():
    raw = "Check with vendor. [Partial dimension found: 14 W; full W x H x D still needed]"
    result = _clean_notes(raw)
    assert "[Partial dimension" not in result
    assert "Check with vendor." in result

def test_clean_notes_removes_row_prefix():
    assert _clean_notes("3 - Verify finish") == "Verify finish"

def test_clean_notes_keeps_human_text():
    assert _clean_notes("Stainless steel finish, confirm with supplier.") == \
        "Stainless steel finish, confirm with supplier."

def test_clean_notes_empty_string():
    assert _clean_notes("") == ""

def test_clean_notes_only_system_tag_returns_empty():
    result = _clean_notes("[Enrichment: no confident source found]")
    assert result == ""


# ── _row_to_programa_dict ─────────────────────────────────────────────────────

def test_scotsman_section_maps_to_product_category():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Section"] == "Appliances"

def test_scotsman_sku_mapped():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["SKU"] == "SCN60PA1SU"

def test_scotsman_model_blank():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Model"] == ""

def test_scotsman_dimensions_direct():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Dimensions"] == "14.875 in W x 22 in D x 33.375 in H"

def test_scotsman_width_parsed():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Width (in)"] == "14.875"

def test_scotsman_height_parsed():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Height (in)"] == "33.375"

def test_scotsman_depth_parsed():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Depth (in)"] == "22"

def test_scotsman_length_blank_when_absent():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Length (in)"] == ""

def test_scotsman_finish_from_finish_color():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Finish"] == "Stainless Steel"

def test_scotsman_color_blank():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Color"] == ""

def test_scotsman_material_extracted_from_notes():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Material"] == "Stainless Steel"

def test_scotsman_notes_cleaned():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert "[Materials:" not in result["Notes"]
    assert "[Enrichment:" not in result["Notes"]
    assert "Verify delivery date." in result["Notes"]

def test_scotsman_location_from_room():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Location"] == "Kitchen"

def test_section_fallback_to_general_when_category_blank():
    from src.programa_export import _row_to_programa_dict
    row = _scotsman_row()
    row["Product Category"] = ""
    result = _row_to_programa_dict(row)
    assert result["Section"] == "General"

def test_material_explicit_field_takes_priority():
    from src.programa_export import _row_to_programa_dict
    row = _scotsman_row()
    row["Material"] = "Cast Iron"
    result = _row_to_programa_dict(row)
    assert result["Material"] == "Cast Iron"

def test_lead_time_blank_when_no_field():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Lead Time"] == ""

def test_lead_time_explicit_field_used():
    from src.programa_export import _row_to_programa_dict
    row = _scotsman_row()
    row["Lead Time"] = "8-10 weeks"
    result = _row_to_programa_dict(row)
    assert result["Lead Time"] == "8-10 weeks"

def test_quantity_coerced_to_int():
    from src.programa_export import _row_to_programa_dict
    result = _row_to_programa_dict(_scotsman_row())
    assert result["Quantity"] == 1

def test_quantity_string_input_coerced():
    from src.programa_export import _row_to_programa_dict
    row = _scotsman_row()
    row["Quantity"] = "3"
    result = _row_to_programa_dict(row)
    assert result["Quantity"] == 3

def test_output_has_all_programa_columns():
    from src.programa_export import _row_to_programa_dict, PROGRAMA_COLUMNS
    result = _row_to_programa_dict(_scotsman_row())
    for col in PROGRAMA_COLUMNS:
        assert col in result, f"Missing column: {col}"


def test_photo_only_export_allows_blank_brand_sku_model_url_dimensions():
    from src.programa_export import build_programa_import_dataframe

    row = {
        "Include": True,
        "Source Type": "Photo",
        "Import Type": "Photo Inventory Upload",
        "photo_only": True,
        "Product Name": "Handmade Ceramic Bowl",
        "Product Category": "Accessories",
        "Image URL": "https://res.cloudinary.com/demo/image/upload/bowl.jpg",
        "Color": "Ivory",
        "Material": "Ceramic",
        "Notes": "Photo-only item; details generated from uploaded image.",
    }

    df = build_programa_import_dataframe([row])
    assert len(df) == 1
    assert df.iloc[0]["Brand"] == ""
    assert df.iloc[0]["SKU"] == ""
    assert df.iloc[0]["Model"] == ""
    assert df.iloc[0]["Product URL"] == ""
    assert df.iloc[0]["Dimensions"] == ""
    assert df.iloc[0]["Section"] == "Decor"
    assert df.iloc[0]["Image URL"].startswith("https://")


def test_photo_only_export_skips_missing_image_url():
    from src.programa_export import build_programa_import_dataframe

    row = {
        "Include": True,
        "Source Type": "Photo",
        "photo_only": True,
        "Product Name": "Handmade Bowl",
        "Product Category": "Accessories",
        "Image URL": "",
    }

    assert build_programa_import_dataframe([row]).empty


def test_photo_only_export_skips_non_https_image_url():
    from src.programa_export import build_programa_import_dataframe, validate_for_export

    row = {
        "Include": True,
        "Source Type": "Photo",
        "photo_only": True,
        "Product Name": "Handmade Bowl",
        "Product Category": "Accessories",
        "Image URL": "http://example.com/bowl.jpg",
    }

    assert build_programa_import_dataframe([row]).empty
    summary = validate_for_export([row])
    assert summary["export_count"] == 0
    assert summary["missing_image_url"] == 1
    assert summary["skipped"][0]["reason"] == "missing or invalid Image URL"


def test_export_blanks_non_https_image_url_for_standard_rows():
    from src.programa_export import build_programa_import_dataframe

    rows = _make_rows([{"Image URL": "http://example.com/image.jpg"}])

    df = build_programa_import_dataframe(rows)

    assert df.iloc[0]["Image URL"] == ""


def test_export_blanks_html_extension_url():
    """https:// URL with .html extension is a product page — must be blanked."""
    from src.programa_export import build_programa_import_dataframe

    rows = _make_rows([{"Image URL": "https://example.com/product/detail.html"}])
    df = build_programa_import_dataframe(rows)
    assert df.iloc[0]["Image URL"] == ""


def test_export_accepts_webp_image_url():
    """WebP images are valid and must pass through to the export."""
    from src.programa_export import build_programa_import_dataframe

    rows = _make_rows([{"Image URL": "https://example.com/photo.webp"}])
    df = build_programa_import_dataframe(rows)
    assert df.iloc[0]["Image URL"] == "https://example.com/photo.webp"


def test_export_accepts_cdn_url_without_extension():
    """CDN image URLs without file extension (Scene7-style) must pass through."""
    from src.programa_export import build_programa_import_dataframe

    rows = _make_rows([{"Image URL": "https://s7d4.scene7.com/is/image/Brand/ModelName"}])
    df = build_programa_import_dataframe(rows)
    assert df.iloc[0]["Image URL"] == "https://s7d4.scene7.com/is/image/Brand/ModelName"


def test_photo_only_bulk_export_contains_public_urls_only():
    from src.photo_inventory import create_photo_only_bulk_rows
    from src.programa_export import build_programa_import_dataframe

    rows = create_photo_only_bulk_rows(
        [
            {"image_filename": "lamp.jpg", "local_image_path": "C:/tmp/lamp.jpg"},
            {"image_filename": "handmade_doll.jpg", "local_image_path": "C:/tmp/handmade_doll.jpg"},
            {"image_filename": "vase.jpg", "local_image_path": "C:/tmp/vase.jpg"},
        ],
        project="",
        room="",
        section="Decor",
        naming_mode="Filename",
        image_urls=[
            "https://cdn.example.com/lamp.jpg",
            "https://cdn.example.com/handmade_doll.jpg",
            "https://cdn.example.com/vase.jpg",
        ],
    )

    df = build_programa_import_dataframe(rows)
    assert list(df["Section"]) == ["Decor", "Decor", "Decor"]
    assert list(df["Product Name"]) == ["lamp", "handmade_doll", "vase"]
    assert list(df["Quantity"]) == [1, 1, 1]
    assert "Local Image Path" not in df.columns
    assert all(url.startswith("https://cdn.example.com/") for url in df["Image URL"])
    for blank_col in ["Brand", "SKU", "Model", "Dimensions", "Width (in)", "Height (in)", "Depth (in)", "Length (in)", "Price", "Supplier", "Product URL", "Material", "Finish", "Color", "Lead Time"]:
        assert list(df[blank_col]) == ["", "", ""]


# ── validate_for_export ───────────────────────────────────────────────────────

from src.programa_export import validate_for_export


def test_validate_counts_export_rows():
    rows = _make_rows([{}, {}, {}])
    result = validate_for_export(rows)
    assert result["export_count"] == 3


def test_validate_skips_rows_missing_product_name():
    rows = _make_rows([{"Product Name": ""}, {"Product Name": "Real Product"}])
    result = validate_for_export(rows)
    assert result["export_count"] == 1
    assert len(result["skipped"]) == 1


def test_validate_skips_excluded_rows():
    rows = _make_rows([{"Include": False}, {}])
    result = validate_for_export(rows)
    assert result["export_count"] == 1


def test_validate_flags_missing_section():
    rows = _make_rows([{"Product Category": ""}, {}])
    result = validate_for_export(rows)
    assert len(result["missing_section"]) == 1
    assert result["missing_section"][0]["product_name"] == "Test Product"


def test_validate_flags_missing_dimensions():
    rows = _make_rows([{"Dimensions": ""}, {}])
    result = validate_for_export(rows)
    assert result["missing_dimensions"] == 1


def test_validate_flags_partial_dimensions():
    rows = _make_rows([{"Dimensions": "12 in W x 10 in H"}, {}])
    result = validate_for_export(rows)
    assert result["missing_dimensions"] == 1


def test_validate_flags_missing_product_url():
    rows = _make_rows([{"Product URL": ""}, {}])
    result = validate_for_export(rows)
    assert result["missing_product_url"] == 1


def test_validate_flags_missing_image_url():
    rows = _make_rows([{"Image URL": ""}, {}])
    result = validate_for_export(rows)
    assert result["missing_image_url"] == 1
    assert result["image_url_present"] == 1
    assert result["image_url_total"] == 2


def test_validate_counts_image_urls_present():
    rows = _make_rows([
        {"Image URL": "https://cdn.example.com/one.jpg"},
        {"Image URL": "https://cdn.example.com/two.jpg"},
        {"Image URL": ""},
    ])
    result = validate_for_export(rows)

    assert result["image_url_present"] == 2
    assert result["image_url_total"] == 3


def test_validate_accepts_dataframe_input():
    df = pd.DataFrame(_make_rows([{}, {}]))
    result = validate_for_export(df)
    assert result["export_count"] == 2


def test_validate_result_keys():
    result = validate_for_export([])
    assert set(result.keys()) == {
        "skipped", "missing_section", "missing_dimensions",
        "missing_product_url", "missing_image_url", "image_url_present",
        "image_url_total", "export_count",
        "unique_sections", "section_counts", "section_equals_product_name",
        "section_too_long", "too_many_unique_sections", "canonical_sections",
        "duplicates_removed", "duplicate_rows_removed",
        "suspicious_dimensions_rejected", "rejected_product_urls",
        "pdf_product_urls", "blank_price_only_rows",
        "missing_model_manufacturer", "phone_email_header_contamination",
        "parsed_rows_count", "export_rows_count",
    }


def test_validate_returns_section_distribution():
    rows = _make_rows([
        {"Product Category": "Seating"},
        {"Product Category": "Lighting"},
        {"Product Category": "Accessories"},
    ])
    result = validate_for_export(rows)
    assert result["unique_sections"] == ["Decor", "Furniture", "Lighting"]
    assert result["section_counts"] == {"Decor": 1, "Furniture": 1, "Lighting": 1}


def test_validate_flags_section_equal_product_name_and_long_raw_section():
    long_section = "This Section Name Is Definitely Too Long"
    rows = _make_rows([
        {"Product Name": "Lamp", "Product Category": "Lamp"},
        {"Product Name": "Long Raw", "Product Category": long_section},
    ])
    result = validate_for_export(rows)
    assert result["section_equals_product_name"][0]["product_name"] == "Lamp"
    assert result["section_too_long"][0]["section"] == long_section
    assert result["too_many_unique_sections"] is False


# ── build_programa_import_dataframe ───────────────────────────────────────────

from src.programa_export import (
    build_programa_import_dataframe,
    build_programa_debug_dataframe,
    export_programa_csv,
    export_programa_xlsx,
    PROGRAMA_COLUMNS,
    CANONICAL_SECTIONS,
    _DEBUG_EXTRA_COLUMNS,
    normalize_section,
)


def test_build_returns_dataframe():
    rows = _make_rows([{}])
    df = build_programa_import_dataframe(rows)
    assert isinstance(df, pd.DataFrame)


def test_build_has_exactly_programa_columns():
    rows = _make_rows([{}])
    df = build_programa_import_dataframe(rows)
    assert list(df.columns) == PROGRAMA_COLUMNS
    assert {"Section", "Product Name", "Image URL", "Quantity", "Notes"}.issubset(df.columns)
    assert "Local Image Path" not in df.columns
    assert "Image Upload Status" not in df.columns
    assert "Confidence Score" not in df.columns


def test_build_excludes_rows_missing_product_name():
    rows = _make_rows([{"Product Name": ""}, {}])
    df = build_programa_import_dataframe(rows)
    assert len(df) == 1


def test_build_excludes_not_included_rows():
    rows = _make_rows([{"Include": False}, {}])
    df = build_programa_import_dataframe(rows)
    assert len(df) == 1


def test_build_scotsman_acceptance_case():
    row = _scotsman_row()
    df = build_programa_import_dataframe([row])
    assert len(df) == 1
    assert df.iloc[0]["Section"] == "Appliances"
    assert df.iloc[0]["Width (in)"] == "14.875"
    assert df.iloc[0]["Height (in)"] == "33.375"
    assert df.iloc[0]["Depth (in)"] == "22"
    assert "[Materials:" not in df.iloc[0]["Notes"]
    assert df.iloc[0]["Material"] == "Stainless Steel"


def test_section_normalization_maps_legacy_categories_to_canonical_sections():
    assert normalize_section("Seating") == "Furniture"
    assert normalize_section("Stone/Tile") == "Flooring"
    assert normalize_section("Accessories") == "Decor"
    assert normalize_section("Totally Custom One-Off") == "General"


def test_export_contains_only_canonical_sections():
    rows = _make_rows([
        {"Product Category": "Seating"},
        {"Product Category": "Stone/Tile"},
        {"Product Category": "Product Name As Section", "Product Name": "Product Name As Section"},
    ])
    df = build_programa_import_dataframe(rows)
    assert set(df["Section"]).issubset(set(CANONICAL_SECTIONS))
    assert list(df["Section"]) == ["Furniture", "Flooring", "General"]


def test_build_accepts_dataframe_input():
    df_in = pd.DataFrame(_make_rows([{}, {}]))
    df_out = build_programa_import_dataframe(df_in)
    assert len(df_out) == 2


def test_build_empty_input_returns_empty_dataframe():
    df = build_programa_import_dataframe([])
    assert isinstance(df, pd.DataFrame)
    assert list(df.columns) == PROGRAMA_COLUMNS
    assert len(df) == 0


def test_build_dedupes_brand_sku_and_keeps_richer_row():
    rows = _make_rows([
        {
            "Product Name": "Lynx Built-In Grill",
            "Brand": "Lynx",
            "Model/SKU": "L42TR",
            "Room": "Outdoor Kitchen",
            "Supplier": "PC Richard",
            "Product URL": "https://lynxgrills.com/products/l42tr",
            "Image URL": "https://lynxgrills.com/images/l42tr.jpg",
        },
        {
            "Product Name": "lynx l42tr",
            "Brand": "lynx",
            "Model/SKU": "L42TR",
            "Room": "",
            "Supplier": "",
            "Product URL": "",
            "Image URL": "",
        },
    ])

    df = build_programa_import_dataframe(rows)
    summary = validate_for_export(rows)

    assert len(df) == 1
    assert df.iloc[0]["Product Name"] == "Lynx Built-In Grill"
    assert df.iloc[0]["Location"] == "Outdoor Kitchen"
    assert summary["duplicate_rows_removed"] == 1


def test_build_preserves_same_sku_in_different_rooms():
    rows = _make_rows([
        {"Brand": "Wolf", "Model/SKU": "WWD30", "Room": "Kitchen"},
        {"Brand": "Wolf", "Model/SKU": "WWD30", "Room": "Pantry"},
    ])

    df = build_programa_import_dataframe(rows)

    assert len(df) == 2
    assert set(df["Location"]) == {"Kitchen", "Pantry"}


def test_build_rejects_implausible_appliance_dimensions():
    rows = _make_rows([
        {
            "Product Category": "Appliances",
            "Brand": "Miele",
            "Model/SKU": "G7986SCVIK20",
            "Dimensions": '2752"H x 4"D x 67.7881"L',
        }
    ])

    df = build_programa_import_dataframe(rows)
    summary = validate_for_export(rows)

    assert df.iloc[0]["Dimensions"] == ""
    assert df.iloc[0]["Height (in)"] == ""
    assert summary["suspicious_dimensions_rejected"][0]["sku"] == "G7986SCVIK20"


def test_build_blanks_sitemap_product_url_and_warns():
    rows = _make_rows([
        {
            "Brand": "Sub-Zero",
            "Model/SKU": "DEC3650RID/R",
            "Product URL": "https://www.subzero-wolf.com/product-sitemap/sub-zero",
        }
    ])

    df = build_programa_import_dataframe(rows)
    summary = validate_for_export(rows)

    assert df.iloc[0]["Product URL"] == ""
    assert summary["rejected_product_urls"][0]["reason"] == "sitemap/category URL rejected"


def test_validate_warns_when_product_url_is_pdf():
    rows = _make_rows([
        {
            "Brand": "Wolf",
            "Model/SKU": "WWD30",
            "Product URL": "https://subzero-wolf.com/specs/wwd30-spec.pdf",
        }
    ])

    summary = validate_for_export(rows)

    assert summary["pdf_product_urls"][0]["sku"] == "WWD30"


def test_validate_reports_parse_qa_issues():
    rows = _make_rows([
        {
            "Include": False,
            "Import Type": "unresolved_charge",
            "Product Name": "",
            "Brand": "",
            "Model/SKU": "",
            "Price": "$285.97",
        },
        {
            "Product Name": "Potential Product",
            "Brand": "",
            "Model/SKU": "",
        },
        {
            "Product Name": "Sub-Zero Refrigerator",
            "Brand": "Sub-Zero",
            "Model/SKU": "631-287-2405",
        },
    ])

    summary = validate_for_export(rows)

    assert summary["blank_price_only_rows"][0]["price"] == "$285.97"
    assert summary["missing_model_manufacturer"][0]["reason"] == "missing manufacturer and model"
    assert summary["phone_email_header_contamination"][0]["reason"] == "phone/email captured as model"
    assert summary["parsed_rows_count"] == 3
    assert summary["export_rows_count"] == 2


# ── build_programa_debug_dataframe ────────────────────────────────────────────

def test_debug_build_has_programa_plus_debug_columns():
    rows = _make_rows([{}])
    df = build_programa_debug_dataframe(rows)
    expected = PROGRAMA_COLUMNS + _DEBUG_EXTRA_COLUMNS
    assert list(df.columns) == expected


# ── export_programa_csv ───────────────────────────────────────────────────────

def test_export_csv_returns_bytes():
    df = build_programa_import_dataframe(_make_rows([{}]))
    result = export_programa_csv(df)
    assert isinstance(result, bytes)


def test_export_csv_contains_header_row():
    df = build_programa_import_dataframe(_make_rows([{}]))
    text = export_programa_csv(df).decode("utf-8")
    assert "Section" in text
    assert "Product Name" in text


def test_export_csv_one_row_per_product():
    rows = _make_rows([{}, {}])
    df = build_programa_import_dataframe(rows)
    text = export_programa_csv(df).decode("utf-8")
    lines = [l for l in text.splitlines() if l.strip()]
    assert len(lines) == 3  # header + 2 rows


# ── export_programa_xlsx ──────────────────────────────────────────────────────

def test_export_xlsx_returns_bytes():
    df = build_programa_import_dataframe(_make_rows([{}]))
    result = export_programa_xlsx(df)
    assert isinstance(result, bytes)


def test_export_xlsx_is_valid_xlsx():
    import openpyxl
    df = build_programa_import_dataframe(_make_rows([{}]))
    xlsx_bytes = export_programa_xlsx(df)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert len(wb.sheetnames) == 1
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, len(PROGRAMA_COLUMNS) + 1)]
    assert headers[0] == "Section"
    assert headers[1] == "Product Name"


def test_export_xlsx_no_merged_cells():
    import openpyxl
    df = build_programa_import_dataframe(_make_rows([{}]))
    wb = openpyxl.load_workbook(io.BytesIO(export_programa_xlsx(df)))
    assert len(wb.active.merged_cells.ranges) == 0


# ── Golden test ───────────────────────────────────────────────────────────────

def test_golden_csv_exact_columns_and_no_nan():
    """
    Golden test: one clean product → CSV has exact column order,
    no extra columns, no NaN strings, and correct values in data row.
    """
    import csv as _csv

    row = _scotsman_row()
    df = build_programa_import_dataframe([row])
    csv_bytes = export_programa_csv(df)
    csv_text = csv_bytes.decode("utf-8")

    # Exact column count and order
    assert len(df.columns) == len(PROGRAMA_COLUMNS)
    assert list(df.columns) == PROGRAMA_COLUMNS

    # No NaN strings anywhere in output
    assert "nan" not in csv_text.lower()

    # Parse and verify structure
    parsed_rows = list(_csv.reader(csv_text.splitlines()))
    assert len(parsed_rows) == 2  # header + 1 data row
    assert parsed_rows[0] == PROGRAMA_COLUMNS

    data = parsed_rows[1]
    assert data[PROGRAMA_COLUMNS.index("Section")] == "Appliances"
    assert data[PROGRAMA_COLUMNS.index("Product Name")] == "Scotsman Icemaker Built-In Pump"
    assert data[PROGRAMA_COLUMNS.index("SKU")] == "SCN60PA1SU"
    assert data[PROGRAMA_COLUMNS.index("Width (in)")] == "14.875"
    assert data[PROGRAMA_COLUMNS.index("Height (in)")] == "33.375"
    assert data[PROGRAMA_COLUMNS.index("Depth (in)")] == "22"
    assert data[PROGRAMA_COLUMNS.index("Material")] == "Stainless Steel"
    assert "[Materials:" not in data[PROGRAMA_COLUMNS.index("Notes")]
