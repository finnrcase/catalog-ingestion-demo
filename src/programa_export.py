"""
Programa import-file export for SCH DesignOps Intake.

Transforms the internal intake DataFrame into a clean CSV/XLSX compatible
with Programa's built-in "Import Products" feature.

Public API
----------
PROGRAMA_COLUMNS : list[str]
    Fixed 21-column output order for the Programa import file.

validate_for_export(rows) -> dict
    Returns a validation summary without modifying data.

build_programa_import_dataframe(rows) -> pd.DataFrame
    Returns a clean export-ready DataFrame (Include=True, Product Name required).

build_programa_debug_dataframe(rows) -> pd.DataFrame
    Same as above plus debug/confidence columns for internal review.

export_programa_csv(df) -> bytes
    Serialize DataFrame to UTF-8 CSV bytes.

export_programa_xlsx(df) -> bytes
    Serialize DataFrame to XLSX bytes (single sheet, no merged cells).
"""

from __future__ import annotations

import datetime
import io
import re
import hashlib
import os
import zipfile
import urllib.parse
from pathlib import Path

import httpx
import pandas as pd
from PIL import Image, ImageOps
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.dimensions import extract_labeled_dimensions, has_complete_3d_dimensions
from src.image_presence import image_filename, local_image_path, row_has_image
from src.notes import remove_notes_row_prefix

# ── Constants ─────────────────────────────────────────────────────────────────

PROGRAMA_COLUMNS: list[str] = [
    "Section",
    "Product Name",
    "Brand",
    "SKU",
    "Model",
    "Dimensions",
    "Width (in)",
    "Height (in)",
    "Depth (in)",
    "Length (in)",
    "Quantity",
    "Price",
    "Supplier",
    "Product URL",
    "Image URL",
    "Finish",
    "Color",
    "Material",
    "Lead Time",
    "Notes",
    "Location",
]

PROGRAMA_IMAGE_FORMAT_NOTE = (
    "Programa's public import docs confirm .xlsx/.xls/.csv uploads and state "
    "that product images must be included on the same row as the product. They "
    "do not document ZIP sidecar filename imports, so the most compatible image "
    "handoff is an XLSX with embedded same-row images; direct public HTTPS image "
    "URLs remain backup references."
)

PROGRAMA_XLSX_WITH_IMAGES_COLUMNS: list[str] = [
    *PROGRAMA_COLUMNS[:15],
    "Product Image",
    "Image Filename",
    "Image Import Status",
    *PROGRAMA_COLUMNS[15:],
]

CANONICAL_SECTIONS: list[str] = [
    "Appliances",
    "Lighting",
    "Plumbing",
    "Cabinetry",
    "Flooring",
    "Furniture",
    "Decor",
    "Hardware",
    "Exterior",
    "General",
]

_DEBUG_EXTRA_COLUMNS: list[str] = [
    "Confidence Score",
    "Source Type",
    "AI Category Confidence",
    "Category Source",
    "Original Image URL",
    "Local Image Path",
    "Image Filename",
    "Image Upload Status",
    "cloudinary_secure_url",
    "cloudinary_url",
    "cloudinary_public_id",
    "cloudinary_width",
    "cloudinary_height",
    "cloudinary_format",
    "cloudinary_bytes",
    "original_image_url",
    "image_confidence",
    "image_source_url",
    "cloudinary_status",
    "cloudinary_error",
    "programa_image_ready",
    "image_upload_status",
    "image_upload_failure_reason",
    "Dimension Source URL",
    "Dimension Confidence",
    "Dimension Source Type",
    "Dimension Lookup Status",
    "Product Width (in)",
    "Product Height (in)",
    "Product Depth (in)",
    "Cutout Dimensions",
    "Cutout Width (in)",
    "Cutout Height (in)",
    "Cutout Depth (in)",
    "Shipping/Package Dimensions",
    "Shipping Width (in)",
    "Shipping Height (in)",
    "Shipping Depth (in)",
    # Phase 1 image recovery internal source metadata
    "_source_pdf_id",
    "_source_page_number",
    "_source_filename",
    # Phase 1 image recovery confidence/evidence
    "image_source",
    "confidence",
    "evidence",
    "needs_image_review",
]

_MATERIAL_TAG_RE = re.compile(r"\[Materials:\s*([^\]]+)\]", re.IGNORECASE)
_SYSTEM_TAG_RE = re.compile(r"\[[^\]]*\]")
_NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")
_SIGNED_URL_QUERY_KEYS = {
    "x-amz-algorithm",
    "x-amz-credential",
    "x-amz-date",
    "x-amz-expires",
    "x-amz-signature",
    "x-amz-security-token",
    "signature",
    "sig",
    "expires",
    "expiry",
    "expires_at",
    "token",
    "access_token",
    "policy",
    "key-pair-id",
}

# Anchor to the repo root regardless of process cwd.
# src/programa_export.py → src/ → repo root (two levels up).
_REPO_ROOT = Path(__file__).resolve().parent.parent


def _tmp_upload_root() -> Path:
    return Path(os.getenv("SCH_TMP_UPLOAD_ROOT", str(_REPO_ROOT / ".tmp" / "uploads"))).expanduser()

_SECTION_ALIASES: dict[str, str] = {
    "": "General",
    "accessories": "Decor",
    "accessory": "Decor",
    "appliance": "Appliances",
    "appliances": "Appliances",
    "art": "Decor",
    "artwork": "Decor",
    "bath": "Plumbing",
    "bath linens": "Decor",
    "bathroom": "Plumbing",
    "bedding": "Decor",
    "bedding linens bath linens": "Decor",
    "beds": "Furniture",
    "beds mattresses": "Furniture",
    "cabinet": "Cabinetry",
    "cabinets": "Cabinetry",
    "cabinetry": "Cabinetry",
    "casework": "Cabinetry",
    "chairs": "Furniture",
    "decor": "Decor",
    "decoration": "Decor",
    "decorative": "Decor",
    "dresser": "Furniture",
    "dressers drawers storage": "Furniture",
    "electrical": "Lighting",
    "exterior": "Exterior",
    "fabric": "Decor",
    "fabrics pillows": "Decor",
    "flooring": "Flooring",
    "furniture": "Furniture",
    "general": "General",
    "gym equipment": "Furniture",
    "hardware": "Hardware",
    "lighting": "Lighting",
    "lights": "Lighting",
    "linen": "Decor",
    "linens": "Decor",
    "millwork": "Cabinetry",
    "mirror": "Decor",
    "mirrors": "Decor",
    "outdoor": "Exterior",
    "paint": "Decor",
    "paint wallpaper": "Decor",
    "pillow": "Decor",
    "pillows": "Decor",
    "plumbing": "Plumbing",
    "rug": "Decor",
    "rugs": "Decor",
    "seating": "Furniture",
    "sofa": "Furniture",
    "stone": "Flooring",
    "stone tile": "Flooring",
    "storage": "Furniture",
    "table": "Furniture",
    "tables": "Furniture",
    "tile": "Flooring",
    "wallpaper": "Decor",
}


# ── Internal helpers ──────────────────────────────────────────────────────────

def _str_val(v) -> str:
    """Safely coerce a cell value to a stripped string, treating None/NaN as blank."""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in {"nan", "none", "null"} else s


def _section_key(value: str) -> str:
    return _NON_ALNUM_RE.sub(" ", value.strip().lower()).strip()


def normalize_section(value: object) -> str:
    """Map any inferred category/section to the controlled Programa section list."""
    raw = _str_val(value)
    key = _section_key(raw)
    if not key:
        return "General"
    canonical_lookup = {_section_key(section): section for section in CANONICAL_SECTIONS}
    if key in canonical_lookup:
        return canonical_lookup[key]
    if key in _SECTION_ALIASES:
        return _SECTION_ALIASES[key]

    # Keyword fallback catches verbose AI labels like "decorative table lamp".
    keyword_map = [
        ("appliance", "Appliances"),
        ("light", "Lighting"),
        ("lamp", "Lighting"),
        ("plumb", "Plumbing"),
        ("sink", "Plumbing"),
        ("faucet", "Plumbing"),
        ("cabinet", "Cabinetry"),
        ("millwork", "Cabinetry"),
        ("floor", "Flooring"),
        ("tile", "Flooring"),
        ("stone", "Flooring"),
        ("chair", "Furniture"),
        ("seat", "Furniture"),
        ("sofa", "Furniture"),
        ("table", "Furniture"),
        ("bed", "Furniture"),
        ("dresser", "Furniture"),
        ("rug", "Decor"),
        ("mirror", "Decor"),
        ("pillow", "Decor"),
        ("decor", "Decor"),
        ("art", "Decor"),
        ("hardware", "Hardware"),
        ("knob", "Hardware"),
        ("pull", "Hardware"),
        ("exterior", "Exterior"),
        ("outdoor", "Exterior"),
    ]
    for needle, section in keyword_map:
        if needle in key:
            return section
    return "General"


def _extract_material_from_notes(notes: str) -> str:
    """Return the value from the first [Materials: ...] tag, or empty string."""
    notes = notes or ""
    m = _MATERIAL_TAG_RE.search(notes)
    return m.group(1).strip() if m else ""


def _clean_notes(notes: str) -> str:
    """Strip all [...] system tags and leading row-number prefixes from notes."""
    text = _SYSTEM_TAG_RE.sub("", notes)
    text = re.sub(r"\s{2,}", " ", text).strip()
    return remove_notes_row_prefix(text)


def _to_row_list(rows) -> list[dict]:
    """Normalize supported row containers to a list of dictionaries."""
    if rows is None:
        return []
    if isinstance(rows, pd.DataFrame):
        return [r.to_dict() for _, r in rows.iterrows()]
    return list(rows)


def _is_included(row: dict) -> bool:
    """True when Include is truthy (True, 1, 'True', or absent)."""
    v = row.get("Include", True)
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    return str(v).strip().lower() not in {"false", "0", "no"}


def _is_photo_only(row: dict) -> bool:
    value = _str_val(row.get("photo_only")).lower()
    return (
        value in {"true", "1", "yes"}
        or _str_val(row.get("Source Type")) == "Photo"
        or _str_val(row.get("Import Type")).lower() in {"photo upload", "photo inventory upload"}
    )


_EXPORT_IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp")
_EXPORT_NON_IMAGE_EXTENSIONS = (".html", ".htm", ".php", ".asp", ".aspx", ".cfm", ".pdf", ".txt", ".xml")


def _is_public_https_image_url(value) -> bool:
    url = _str_val(value).lower()
    if not url.startswith("https://"):
        return False
    path = url.split("?")[0].split("#")[0]
    if path.endswith(_EXPORT_IMAGE_EXTENSIONS):
        return True
    if path.endswith(_EXPORT_NON_IMAGE_EXTENSIONS):
        return False
    # CDN heuristic: no file extension in the last path segment → accept.
    # Enrichment already validated these URLs via HEAD content-type check.
    last_segment = path.rstrip("/").rsplit("/", 1)[-1]
    return "." not in last_segment


def _has_temporary_or_signed_query(url: str) -> bool:
    parsed = urllib.parse.urlparse(url)
    query_keys = {key.lower() for key, _ in urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)}
    return bool(query_keys & _SIGNED_URL_QUERY_KEYS)


def validate_public_image_url(url: str | None, timeout: float = 6.0) -> dict:
    """
    Strict Programa image-URL validation.

    Programa's importer may map direct public image URLs, but product-page URLs,
    local paths, signed temporary links, and inaccessible links are not safe
    import image values. This probe verifies the final URL is HTTPS, reachable,
    and serves image/* content.
    """
    raw = _str_val(url)
    result = {
        "ok": False,
        "url": raw,
        "final_url": "",
        "status_code": "",
        "content_type": "",
        "reason": "",
        "is_jpeg": False,
    }
    if not raw:
        result["reason"] = "missing_url"
        return result
    parsed = urllib.parse.urlparse(raw)
    if parsed.scheme.lower() != "https":
        result["reason"] = "not_https"
        return result
    if _has_temporary_or_signed_query(raw):
        result["reason"] = "signed_or_temporary_url"
        return result
    if not _is_public_https_image_url(raw):
        result["reason"] = "not_direct_image_url"
        return result

    headers = {"User-Agent": "Mozilla/5.0 (compatible; SCH-Intake/1.0)"}
    response = None
    try:
        response = httpx.head(raw, headers=headers, timeout=timeout, follow_redirects=True)
        if response.status_code in {403, 405} or not response.headers.get("content-type"):
            response = httpx.get(
                raw,
                headers={**headers, "Range": "bytes=0-2047"},
                timeout=timeout,
                follow_redirects=True,
            )
    except Exception as exc:
        result["reason"] = f"request_failed:{exc}"
        return result

    final_url = str(response.url)
    content_type = response.headers.get("content-type", "").lower().split(";")[0].strip()
    result["final_url"] = final_url
    result["status_code"] = response.status_code
    result["content_type"] = content_type

    final_parsed = urllib.parse.urlparse(final_url)
    if final_parsed.scheme.lower() != "https":
        result["reason"] = "redirected_to_non_https"
        return result
    if _has_temporary_or_signed_query(final_url):
        result["reason"] = "redirected_to_signed_or_temporary_url"
        return result
    if response.status_code not in {200, 206}:
        result["reason"] = f"bad_status:{response.status_code}"
        return result
    if not content_type.startswith("image/"):
        result["reason"] = f"not_image_content_type:{content_type or 'missing'}"
        return result
    result["ok"] = True
    result["reason"] = "ok"
    path = final_parsed.path.lower()
    result["is_jpeg"] = content_type in {"image/jpeg", "image/jpg"} or path.endswith((".jpg", ".jpeg"))
    return result


def _programa_image_url(row: dict, *, validate_urls: bool = False) -> str:
    raw_image_url = _str_val(row.get("Image URL"))
    if not raw_image_url:
        return ""
    url = _str_val(row.get("cloudinary_secure_url") or row.get("cloudinary_url") or raw_image_url)
    if not _is_public_https_image_url(url):
        return ""
    if validate_urls and not validate_public_image_url(url).get("ok"):
        return ""
    return url


def _validate_local_path(path_str: str, session_id: str | None) -> tuple[bool, str]:
    """
    Verify that `path_str` is a real .jpg/.jpeg under
    .tmp/uploads/{session_id}/images/ at the repo root.
    Returns (ok, reason). reason is the empty string when ok.
    """
    if not session_id:
        return False, "no_session_id"
    try:
        p = Path(path_str).resolve()
    except Exception:
        return False, "invalid_path"
    if not p.exists():
        return False, "file_not_found"
    if p.suffix.lower() not in (".jpg", ".jpeg"):
        return False, "wrong_extension"
    if p.stat().st_size <= 0:
        return False, "empty_file"
    allowed_root = _tmp_upload_root() / session_id / "images"
    try:
        p.relative_to(allowed_root)
    except ValueError:
        return False, "path_outside_session_dir"
    return True, ""


def _is_exportable(row: dict) -> bool:
    if not (_is_included(row) and _str_val(row.get("Product Name"))):
        return False
    if _is_photo_only(row):
        return row_has_image(row)
    return True


def _quantity_value(value):
    text = _str_val(value)
    if not text:
        return ""
    try:
        numeric = float(text)
        return int(numeric) if numeric.is_integer() else numeric
    except ValueError:
        return text


def _row_to_programa_dict(row: dict, *, validate_image_urls: bool = False) -> dict:
    """Map one internal intake row to Programa's import columns."""
    dimensions = _str_val(row.get("Dimensions"))
    parts = extract_labeled_dimensions(dimensions)
    finish_color = _str_val(row.get("Finish / Color"))
    color = _str_val(row.get("Color"))
    material = _str_val(row.get("Material")) or _extract_material_from_notes(_str_val(row.get("Notes")))
    canonical_model = _str_val(row.get("Model"))

    return {
        "Section": normalize_section(row.get("Product Category") or row.get("Section")),
        "Product Name": _str_val(row.get("Product Name")),
        "Brand": _str_val(row.get("Brand")),
        "SKU": _str_val(row.get("Model/SKU")),
        "Model": canonical_model,
        "Dimensions": dimensions,
        "Width (in)": _str_val(parts.get("width")),
        "Height (in)": _str_val(parts.get("height")),
        "Depth (in)": _str_val(parts.get("depth")),
        "Length (in)": _str_val(parts.get("length")),
        "Quantity": _quantity_value(row.get("Quantity")),
        "Price": _str_val(row.get("Price")),
        "Supplier": _str_val(row.get("Supplier")),
        "Product URL": _str_val(row.get("Product URL")),
        "Image URL": _programa_image_url(row, validate_urls=validate_image_urls),
        "Finish": finish_color,
        "Color": color,
        "Material": material,
        "Lead Time": _str_val(row.get("Lead Time")),
        "Notes": _clean_notes(_str_val(row.get("Notes"))),
        "Location": _str_val(row.get("Room")),
    }


def validate_for_export(rows) -> dict:
    """
    Return a validation summary dict without modifying source rows.

    Photo-only rows require Product Name, Product Category/Section, and Image URL.
    Standard rows require Product Name to appear in the export.
    """
    row_list = _to_row_list(rows)
    included = [r for r in row_list if _is_included(r)]

    skipped: list[dict] = []
    missing_section: list[dict] = []
    missing_dimensions = 0
    missing_product_url = 0
    missing_image_url = 0
    image_url_present = 0
    image_url_total = 0
    export_count = 0
    section_counts: dict[str, int] = {}
    section_equals_product_name: list[dict] = []
    section_too_long: list[dict] = []

    for i, row in enumerate(included):
        name = _str_val(row.get("Product Name"))
        if not name:
            skipped.append({"index": i, "product_name": "(no name)"})
            continue
        image_url_total += 1
        if row_has_image(row):
            image_url_present += 1
        raw_section = _str_val(row.get("Product Category") or row.get("Section"))
        section = normalize_section(raw_section)
        if _is_photo_only(row) and not row_has_image(row):
            missing_image_url += 1
            skipped.append({"index": i, "product_name": name, "reason": "missing image"})
            continue

        export_count += 1
        section_counts[section] = section_counts.get(section, 0) + 1
        if not raw_section:
            missing_section.append({"index": i, "product_name": name})
        if raw_section and _section_key(raw_section) == _section_key(name):
            section_equals_product_name.append({"index": i, "product_name": name, "section": raw_section})
        if len(raw_section) > 30:
            section_too_long.append({"index": i, "product_name": name, "section": raw_section})
        if not has_complete_3d_dimensions(_str_val(row.get("Dimensions"))):
            missing_dimensions += 1
        if not _str_val(row.get("Product URL")):
            missing_product_url += 1
        if not row_has_image(row):
            missing_image_url += 1

    return {
        "skipped": skipped,
        "missing_section": missing_section,
        "missing_dimensions": missing_dimensions,
        "missing_product_url": missing_product_url,
        "missing_image_url": missing_image_url,
        "image_url_present": image_url_present,
        "image_url_total": image_url_total,
        "export_count": export_count,
        "unique_sections": sorted(section_counts),
        "section_counts": dict(sorted(section_counts.items())),
        "section_equals_product_name": section_equals_product_name,
        "section_too_long": section_too_long,
        "too_many_unique_sections": len(section_counts) > len(CANONICAL_SECTIONS),
        "canonical_sections": CANONICAL_SECTIONS,
    }


def build_programa_import_dataframe(rows, *, validate_image_urls: bool = False) -> pd.DataFrame:
    """Transform included intake rows with Product Name into a Programa import DataFrame."""
    records = [
        _row_to_programa_dict(r, validate_image_urls=validate_image_urls)
        for r in _to_row_list(rows)
        if _is_exportable(r)
    ]
    if not records:
        return pd.DataFrame(columns=PROGRAMA_COLUMNS)
    return pd.DataFrame(records, columns=PROGRAMA_COLUMNS)


def build_programa_debug_dataframe(rows) -> pd.DataFrame:
    """Build Programa import rows with debug/confidence columns appended.

    Internal ``_source_*`` columns preserve their original Python type (int,
    str, …) so callers can distinguish unset (None) from numeric zero.
    All other debug columns are coerced to stripped strings via ``_str_val``.
    """
    records = []
    for r in _to_row_list(rows):
        if not _is_exportable(r):
            continue
        mapped = _row_to_programa_dict(r)
        for col in _DEBUG_EXTRA_COLUMNS:
            if col.startswith("_source_"):
                # Preserve raw type (int page numbers, etc.)
                mapped[col] = r.get(col)
            else:
                mapped[col] = _str_val(r.get(col))
        records.append(mapped)
    if not records:
        return pd.DataFrame(columns=PROGRAMA_COLUMNS + _DEBUG_EXTRA_COLUMNS)
    return pd.DataFrame(records, columns=PROGRAMA_COLUMNS + _DEBUG_EXTRA_COLUMNS)


def export_programa_csv(df: pd.DataFrame) -> bytes:
    """Serialize a Programa import DataFrame to UTF-8 CSV bytes."""
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8")


def generate_programa_export_filename(
    rows,
    *,
    extension: str = "csv",
    today: str | None = None,
    kind: str = "import",
    existing_filenames: set[str] | list[str] | tuple[str, ...] | None = None,
    max_length: int = 120,
) -> str:
    """Build a descriptive, filesystem-safe Programa export filename.

    The export data shape stays unchanged; this only improves the attachment
    filename used by API responses and browser downloads.
    """
    row_list = _to_row_list(rows)
    date_str = today or datetime.date.today().isoformat()
    ext = extension.lower().lstrip(".") or "csv"
    project = _first_non_empty(row_list, ("Project", "Project Name", "project", "project_name"))
    room = _first_non_empty(row_list, ("Room", "Location", "room", "location"))
    supplier = _first_non_empty(row_list, ("Supplier", "Vendor", "supplier", "vendor"))
    category = _dominant_value(row_list, ("Product Category", "Section", "Category", "product_category", "section"))

    if not project and room:
        project = f"SCH {room}"
    if not project:
        project = "Untitled Project"

    parts = [project]
    if supplier:
        parts.append(supplier)
    if category:
        parts.append(category)
    label = "Programa Export" if kind == "zip" else "Programa Import"
    parts.append(label)

    stem_parts = [part for part in (_filename_part(part) for part in parts) if part]
    clean_date = re.sub(r"[^0-9-]+", "", date_str) or datetime.date.today().isoformat()
    stem_parts.append(clean_date)
    stem = "_".join(stem_parts)
    filename = _trim_filename(f"{stem}.{ext}", max_length=max_length)
    return _dedupe_export_filename(filename, existing_filenames or set(), max_length=max_length)


def _first_non_empty(rows: list[dict], fields: tuple[str, ...]) -> str:
    for row in rows:
        for field in fields:
            value = _str_val(row.get(field))
            if value:
                return value
    return ""


def _dominant_value(rows: list[dict], fields: tuple[str, ...]) -> str:
    counts: dict[str, tuple[str, int]] = {}
    for row in rows:
        for field in fields:
            value = _str_val(row.get(field))
            if not value:
                continue
            key = re.sub(r"\s+", " ", value.lower()).strip()
            if key:
                counts[key] = (value, counts.get(key, (value, 0))[1] + 1)
                break
    if not counts:
        return ""
    return sorted(counts.values(), key=lambda item: (-item[1], item[0].lower()))[0][0]


def _filename_part(value: object) -> str:
    text = _str_val(value)
    text = re.sub(r"[^\w\s-]+", "", text, flags=re.UNICODE)
    text = re.sub(r"[-\s]+", "_", text).strip("_")
    return text or ""


def _trim_filename(filename: str, *, max_length: int = 120) -> str:
    if len(filename) <= max_length:
        return filename
    stem, dot, ext = filename.rpartition(".")
    if not dot:
        return filename[:max_length].rstrip("_")
    keep = max(1, max_length - len(ext) - 1)
    return f"{stem[:keep].rstrip('_')}.{ext}"


def _dedupe_export_filename(
    filename: str,
    existing_filenames: set[str] | list[str] | tuple[str, ...],
    *,
    max_length: int = 120,
) -> str:
    existing = {str(name).lower() for name in existing_filenames}
    if filename.lower() not in existing:
        return filename
    stem, dot, ext = filename.rpartition(".")
    if not dot:
        stem, ext = filename, ""
    version = 2
    while True:
        suffix = f"_v{version}"
        candidate_stem = stem
        if len(candidate_stem) + len(suffix) + (1 + len(ext) if ext else 0) > max_length:
            trim_to = max_length - len(suffix) - (1 + len(ext) if ext else 0)
            candidate_stem = candidate_stem[:trim_to].rstrip("_")
        candidate = f"{candidate_stem}{suffix}.{ext}" if ext else f"{candidate_stem}{suffix}"
        if candidate.lower() not in existing:
            return candidate
        version += 1


def export_programa_xlsx(df: pd.DataFrame) -> bytes:
    """Serialize a Programa import DataFrame to native XLSX workbook bytes."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Programa Import")
        ws = writer.book["Programa Import"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill(fill_type="solid", fgColor="F7F4EF")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Keep the export lightweight while making it comfortable to review in Excel.
        for idx, column_name in enumerate(df.columns, start=1):
            values = [_str_val(column_name)]
            if not df.empty:
                values.extend(_str_val(value) for value in df.iloc[:, idx - 1].tolist()[:200])
            max_len = max((len(value) for value in values), default=10)
            ws.column_dimensions[get_column_letter(idx)].width = min(60, max(10, max_len + 2))
    return buf.getvalue()


def _local_jpg_status(row: dict, session_id: str | None) -> tuple[bool, str, str]:
    path = local_image_path(row)
    if not path:
        return False, "", "missing_local_image_path"
    if session_id:
        ok, reason = _validate_local_path(path, session_id)
        return ok, path if ok else "", reason
    try:
        p = Path(path).resolve()
    except Exception:
        return False, "", "invalid_path"
    if not p.exists():
        return False, "", "file_not_found"
    if p.suffix.lower() not in (".jpg", ".jpeg"):
        return False, "", "wrong_extension"
    if p.stat().st_size <= 0:
        return False, "", "empty_file"
    return True, str(p), "ok"


def _normalise_image_bytes(raw: bytes, max_size: tuple[int, int] | None = None) -> bytes:
    with Image.open(io.BytesIO(raw)) as img:
        img = ImageOps.exif_transpose(img)
        if img.mode != "RGB":
            img = img.convert("RGB")
        if max_size:
            img.thumbnail(max_size)
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=85, optimize=True)
        return out.getvalue()


def _image_bytes_from_local(path: str) -> bytes:
    return _normalise_image_bytes(Path(path).read_bytes())


def _image_bytes_from_manual(raw: bytes) -> bytes:
    return _normalise_image_bytes(raw)


def _image_bytes_from_url(url: str, row: dict) -> tuple[bytes, str, str]:
    from src.image_assets import download_and_convert_image

    validation = validate_public_image_url(url)
    if not validation.get("ok"):
        return b"", "", str(validation.get("reason") or "invalid_url")
    result = download_and_convert_image(
        str(validation.get("final_url") or url),
        brand=_str_val(row.get("Brand")),
        model_sku=_str_val(row.get("Model/SKU")),
        product_name=_str_val(row.get("Product Name")),
    )
    if result.get("image_status") != "downloaded" or not result.get("jpeg_bytes"):
        return b"", "", str(result.get("error") or result.get("image_status") or "download_failed")
    return result["jpeg_bytes"], result["local_image_filename"], "ok"


def _dedupe_filename(filename: str, seen: dict[str, int], salt: str = "") -> str:
    if filename not in seen:
        seen[filename] = 1
        return filename
    seen[filename] += 1
    digest = hashlib.sha1(f"{filename}:{salt}:{seen[filename]}".encode("utf-8")).hexdigest()[:8]
    if "." in filename:
        stem, ext = filename.rsplit(".", 1)
        return f"{stem}_{digest}.{ext}"
    return f"{filename}_{digest}"


def _base_image_filename(row: dict) -> str:
    from src.image_assets import build_image_filename

    return build_image_filename(
        brand=_str_val(row.get("Brand")),
        model_sku=_str_val(row.get("Model/SKU")),
        product_name=_str_val(row.get("Product Name")),
    )


def _row_image_confidence(row: dict) -> str:
    return _str_val(row.get("confidence")).upper()


def _low_confidence_disallowed(row: dict, include_low_confidence_images: bool) -> bool:
    confidence = _row_image_confidence(row)
    source = _str_val(row.get("image_source")).lower()
    if confidence == "LOW":
        return True
    if confidence == "MEDIUM" and source != "manual_upload":
        return True
    return False


def _resolve_programa_row_image(
    row: dict,
    *,
    row_position: int,
    manual_images: dict | None = None,
    session_id: str | None = None,
    include_low_confidence_images: bool = False,
    seen_filenames: dict[str, int] | None = None,
    download_remote_images: bool = True,
) -> dict:
    """
    Resolve a row image for Programa-compatible export.

    Most compatible path: embed JPEG bytes in the XLSX row. Backup path:
    include a strict direct public HTTPS image URL. Local filesystem paths are
    never written to CSV/XLSX cells.
    """
    manual_images = manual_images or {}
    seen_filenames = seen_filenames if seen_filenames is not None else {}
    status = {
        "bytes": b"",
        "filename": "",
        "image_url": "",
        "status": "manual_upload_required",
        "reason": "",
        "source": "",
        "ready_for_programa": False,
        "local_jpg_available": False,
        "valid_url": False,
        "needs_manual_upload": True,
    }

    if _low_confidence_disallowed(row, include_low_confidence_images):
        status["status"] = "review_confidence_skipped"
        status["reason"] = f"{_row_image_confidence(row).lower()}_confidence_requires_review"
        return status

    local_ok, local_path, local_reason = _local_jpg_status(row, session_id)
    status["local_jpg_available"] = local_ok
    if local_ok:
        try:
            filename = _dedupe_filename(image_filename(row) or _base_image_filename(row), seen_filenames, salt=str(row_position))
            status.update(
                {
                    "bytes": _image_bytes_from_local(local_path),
                    "filename": filename,
                    "status": "embedded_local_jpg",
                    "reason": "ok",
                    "source": "local_image_path",
                    "ready_for_programa": True,
                    "needs_manual_upload": False,
                }
            )
            return status
        except Exception as exc:
            status["reason"] = f"local_image_read_failed:{exc}"
    elif local_reason != "missing_local_image_path":
        status["reason"] = local_reason

    if manual_images.get(row_position):
        try:
            filename = _dedupe_filename(_base_image_filename(row), seen_filenames, salt=f"manual:{row_position}")
            status.update(
                {
                    "bytes": _image_bytes_from_manual(manual_images[row_position]),
                    "filename": filename,
                    "status": "embedded_manual_upload",
                    "reason": "ok",
                    "source": "manual_upload",
                    "ready_for_programa": True,
                    "needs_manual_upload": False,
                }
            )
            return status
        except Exception as exc:
            status["reason"] = f"manual_image_conversion_failed:{exc}"

    image_url = _programa_image_url(row)
    if image_url:
        validation = validate_public_image_url(image_url)
        status["valid_url"] = bool(validation.get("ok"))
        if validation.get("ok"):
            status["image_url"] = str(validation.get("final_url") or image_url)
            if not download_remote_images:
                status.update(
                    {
                        "status": "valid_image_url",
                        "reason": "ok",
                        "source": "image_url",
                        "ready_for_programa": True,
                        "needs_manual_upload": False,
                    }
                )
                return status
            jpeg, filename, reason = _image_bytes_from_url(status["image_url"], row)
            if jpeg:
                filename = _dedupe_filename(filename or _base_image_filename(row), seen_filenames, salt=f"url:{row_position}")
                status.update(
                    {
                        "bytes": jpeg,
                        "filename": filename,
                        "status": "embedded_remote_url",
                        "reason": "ok",
                        "source": "image_url",
                        "ready_for_programa": True,
                        "needs_manual_upload": False,
                    }
                )
                return status
            status.update(
                {
                    "status": "valid_image_url_only",
                    "reason": reason,
                    "source": "image_url",
                    "ready_for_programa": True,
                    "needs_manual_upload": False,
                }
            )
            return status
        status["status"] = "invalid_image_url"
        status["reason"] = str(validation.get("reason") or "invalid_url")

    if not status["reason"]:
        status["reason"] = "missing_image"
    return status


def validate_programa_image_compatibility(
    rows,
    *,
    manual_images: dict | None = None,
    session_id: str | None = None,
    include_low_confidence_images: bool = False,
) -> dict:
    """Return per-row and aggregate image-export compatibility diagnostics."""
    manual_images = manual_images or {}
    row_list = [r for r in _to_row_list(rows) if _is_exportable(r)]
    seen: dict[str, int] = {}
    records: list[dict] = []
    for i, row in enumerate(row_list):
        image = _resolve_programa_row_image(
            row,
            row_position=i,
            manual_images=manual_images,
            session_id=session_id,
            include_low_confidence_images=include_low_confidence_images,
            seen_filenames=seen,
            download_remote_images=False,
        )
        records.append(
            {
                "Product Name": _str_val(row.get("Product Name")),
                "Brand": _str_val(row.get("Brand")),
                "SKU/Model": _str_val(row.get("Model/SKU")),
                "Image URL": image.get("image_url") or _programa_image_url(row),
                "Image Filename": image.get("filename", ""),
                "Image Import Status": image.get("status", ""),
                "Reason": image.get("reason", ""),
                "Local JPG Available": bool(image.get("local_jpg_available")),
                "Valid Public Image URL": bool(image.get("valid_url")),
                "Ready for Programa Image Export": bool(image.get("ready_for_programa")),
                "Needs Manual Upload": bool(image.get("needs_manual_upload")),
            }
        )

    ready = sum(1 for r in records if r["Ready for Programa Image Export"])
    missing = sum(1 for r in records if r["Image Import Status"] in {"manual_upload_required"})
    invalid_urls = sum(1 for r in records if r["Image Import Status"] == "invalid_image_url")
    local = sum(1 for r in records if r["Local JPG Available"])
    valid_urls = sum(1 for r in records if r["Valid Public Image URL"])
    manual_needed = sum(1 for r in records if r["Needs Manual Upload"])
    return {
        "note": PROGRAMA_IMAGE_FORMAT_NOTE,
        "total_rows": len(records),
        "ready_for_programa": ready,
        "missing_images": missing,
        "invalid_urls": invalid_urls,
        "local_jpg_available": local,
        "valid_public_image_urls": valid_urls,
        "manual_upload_needed": manual_needed,
        "rows": records,
    }


def build_programa_image_compatibility_dataframe(summary: dict) -> pd.DataFrame:
    columns = [
        "Product Name",
        "Brand",
        "SKU/Model",
        "Image URL",
        "Image Filename",
        "Image Import Status",
        "Reason",
        "Local JPG Available",
        "Valid Public Image URL",
        "Ready for Programa Image Export",
        "Needs Manual Upload",
    ]
    rows = summary.get("rows") or []
    return pd.DataFrame(rows, columns=columns)


def export_programa_xlsx_with_images(
    rows,
    *,
    manual_images: dict | None = None,
    session_id: str | None = None,
    include_low_confidence_images: bool = False,
) -> bytes:
    """
    Generate Programa's most compatible image import file.

    Programa's public docs say spreadsheet images should be on the same row as
    the product. This writer embeds normalized JPEG thumbnails in a dedicated
    Product Image column, while preserving direct Image URL / Image Filename
    backup references and avoiding any local filesystem paths in cells.
    """
    manual_images = manual_images or {}
    row_list = [r for r in _to_row_list(rows) if _is_exportable(r)]
    seen: dict[str, int] = {}
    records: list[dict] = []
    embedded: list[tuple[int, bytes]] = []

    for i, row in enumerate(row_list):
        mapped = _row_to_programa_dict(row, validate_image_urls=True)
        image = _resolve_programa_row_image(
            row,
            row_position=i,
            manual_images=manual_images,
            session_id=session_id,
            include_low_confidence_images=include_low_confidence_images,
            seen_filenames=seen,
        )
        mapped["Product Image"] = "embedded" if image.get("bytes") else ""
        mapped["Image Filename"] = image.get("filename", "")
        mapped["Image Import Status"] = image.get("status", "")
        if image.get("image_url"):
            mapped["Image URL"] = image["image_url"]
        records.append(mapped)
        if image.get("bytes"):
            embedded.append((len(records) + 1, image["bytes"]))

    df = pd.DataFrame(records, columns=PROGRAMA_XLSX_WITH_IMAGES_COLUMNS)
    xlsx = export_programa_xlsx(df)
    wb = load_workbook(io.BytesIO(xlsx))
    ws = wb.active
    ws.title = "Programa Import"
    image_col_idx = PROGRAMA_XLSX_WITH_IMAGES_COLUMNS.index("Product Image") + 1
    image_col_letter = ws.cell(row=1, column=image_col_idx).column_letter
    ws.column_dimensions[image_col_letter].width = 18
    for row_idx in range(2, len(records) + 2):
        ws.row_dimensions[row_idx].height = 78
        ws.cell(row=row_idx, column=image_col_idx).alignment = Alignment(horizontal="center", vertical="center")

    image_streams: list[io.BytesIO] = []
    for row_idx, raw in embedded:
        thumb = _normalise_image_bytes(raw, max_size=(96, 96))
        stream = io.BytesIO(thumb)
        image_streams.append(stream)
        xl_img = OpenPyXLImage(stream)
        xl_img.width = 96
        xl_img.height = 96
        ws.add_image(xl_img, ws.cell(row=row_idx, column=image_col_idx).coordinate)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


_MANIFEST_COLUMNS: list[str] = [
    "Product Name",
    "Brand",
    "SKU/Model",
    "Product URL",
    "Original Image URL",
    "Local Image Filename",
    "Image Status",
    "Image Source",
    "Confidence",
    "Evidence",
    "Needs Image Review",
    "Error",
]

_MANUAL_GUIDE_COLUMNS: list[str] = [
    "Product Name",
    "Brand",
    "SKU/Model",
    "Product URL",
    "Image Filename",
    "Image Folder Path",
    "Image Source",
    "Confidence",
    "Needs Image Review",
    "Manual Upload Action",
]


def _manual_image_readme() -> str:
    return (
        "SCH DesignOps Programa Image Package\n"
        "====================================\n\n"
        "Programa CSV/XLSX imports may not automatically create product photos from "
        "Image URL or Image Filename columns. Treat those columns as references unless "
        "your Programa import screen explicitly confirms image support.\n\n"
        "Recommended workflow:\n"
        "1. Import product data with programa_import.csv or programa_import.xlsx.\n"
        "2. Open manual_image_upload_guide.csv to match each product to its JPG.\n"
        "3. In Programa, attach the matching file from the images/ folder to each product.\n"
        "4. Use manifest.csv for source, confidence, and troubleshooting details.\n\n"
        "Image notes:\n"
        "- All packaged images are normalized JPG files.\n"
        "- HIGH and MEDIUM confidence images are included by default.\n"
        "- LOW confidence images are excluded unless explicitly enabled before export.\n"
        "- Rows without an image are listed in the guide and manifest for review.\n"
    )


def _manual_upload_action(manifest_row: dict) -> str:
    status = _str_val(manifest_row.get("Image Status"))
    filename = _str_val(manifest_row.get("Local Image Filename"))
    if filename:
        return "Attach images/{filename} to this product in Programa.".format(filename=filename)
    if status == "low_confidence_skipped":
        return "Review low-confidence image before attaching, or re-export with low-confidence images enabled."
    if status == "missing_image_url":
        return "No image was packaged; upload or recover an image before attaching."
    if status in {"fetch_error", "conversion_error", "invalid_local_path"}:
        return "Image was found but could not be packaged; check manifest.csv error and upload manually."
    return "Review this row before attaching an image."


def _manual_guide_dataframe(manifest_rows: list[dict]) -> pd.DataFrame:
    records = []
    for row in manifest_rows:
        filename = _str_val(row.get("Local Image Filename"))
        records.append(
            {
                "Product Name": _str_val(row.get("Product Name")),
                "Brand": _str_val(row.get("Brand")),
                "SKU/Model": _str_val(row.get("SKU/Model")),
                "Product URL": _str_val(row.get("Product URL")),
                "Image Filename": filename,
                "Image Folder Path": f"images/{filename}" if filename else "",
                "Image Source": _str_val(row.get("Image Source")),
                "Confidence": _str_val(row.get("Confidence")),
                "Needs Image Review": _str_val(row.get("Needs Image Review")),
                "Manual Upload Action": _manual_upload_action(row),
            }
        )
    return pd.DataFrame(records, columns=_MANUAL_GUIDE_COLUMNS)


def export_programa_zip(
    rows,
    include_images: bool = True,
    manual_images: dict | None = None,
    session_id: str | None = None,
    include_low_confidence_images: bool = False,
) -> bytes:
    """
    Build a ZIP archive for the Programa export.

    The CSV/XLSX keep Image URL and Image Filename as helper/reference fields,
    but the reliable image handoff is the images/ folder plus
    manual_image_upload_guide.csv. Do not assume Programa imports product
    photos from CSV text fields unless that behavior is verified in the target
    account.

    Image-resolution priority per row:
      1. local_image_path (validated under .tmp/uploads/{session_id}/images/, .jpg)
      2. manual_images[i] bytes
      3. row's Image URL (download via download_and_convert_image)
      4. otherwise: status "missing_image_url"

    LOW-confidence rows are only copied to images/ when
    include_low_confidence_images=True; manifest records the skip otherwise.
    """
    from src.image_assets import download_and_convert_image as _download_convert, build_image_filename

    manual_images = manual_images or {}
    row_list = _to_row_list(rows)

    seen_filenames: dict[str, int] = {}

    def _unique_filename(filename: str, salt: str = "") -> str:
        if filename not in seen_filenames:
            seen_filenames[filename] = 1
            return filename
        seen_filenames[filename] += 1
        digest = hashlib.sha1(f"{filename}:{salt}:{seen_filenames[filename]}".encode("utf-8")).hexdigest()[:8]
        if "." in filename:
            stem, ext = filename.rsplit(".", 1)
            return f"{stem}_{digest}.{ext}"
        return f"{filename}_{digest}"

    def _normalize_manual_jpeg(raw: bytes) -> bytes:
        with Image.open(io.BytesIO(raw)) as img:
            img = ImageOps.exif_transpose(img)
            if img.mode != "RGB":
                img = img.convert("RGB")
            out = io.BytesIO()
            img.save(out, format="JPEG", quality=85, optimize=True)
            return out.getvalue()

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        manifest_rows: list[dict] = []
        export_rows: list[dict] = []
        export_image_filenames: list[str] = []
        for i, r in enumerate(row_list):
            if not _is_exportable(r):
                continue
            export_row = _row_to_programa_dict(r)
            export_rows.append(export_row)
            export_image_filename = ""

            image_url = _str_val(r.get("Image URL"))
            brand = _str_val(r.get("Brand"))
            sku = _str_val(r.get("Model/SKU"))
            product_name = _str_val(r.get("Product Name"))
            confidence = _str_val(r.get("confidence")).upper()
            local_path = local_image_path(r)
            local_filename = image_filename(r)

            # needs_image_review is stored as string "True"/"False" by image_recovery
            # (Task 7 contract — pandas string dtype constraint). Compare against the
            # string explicitly; do not call bool() on it (bool("False") is True).
            _nir_raw = str(r.get("needs_image_review", "")).strip().lower()

            manifest_row: dict = {
                "Product Name": product_name,
                "Brand": brand,
                "SKU/Model": sku,
                "Product URL": _str_val(r.get("Product URL")),
                "Original Image URL": image_url,
                "Local Image Filename": "",
                "Image Status": "missing_image_url",
                "Image Source": _str_val(r.get("image_source")),
                "Confidence": confidence,
                "Evidence": _str_val(r.get("evidence")),
                "Needs Image Review": "false" if _nir_raw == "false" else "true",
                "Error": "",
            }

            if not include_images:
                manifest_rows.append(manifest_row)
                export_image_filenames.append("")
                continue

            if confidence == "LOW" and not include_low_confidence_images:
                manifest_row["Image Status"] = "low_confidence_skipped"
                manifest_rows.append(manifest_row)
                export_image_filenames.append("")
                continue

            wrote_image = False

            # 1) local_image_path
            if local_path:
                ok, reason = _validate_local_path(local_path, session_id)
                if ok:
                    filename = _unique_filename(
                        local_filename or build_image_filename(brand, sku, product_name),
                        salt=f"{i}:{product_name}",
                    )
                    manifest_row["Local Image Filename"] = filename
                    manifest_row["Image Status"] = "downloaded"
                    zf.writestr(f"images/{filename}", Path(local_path).read_bytes())
                    export_image_filename = filename
                    wrote_image = True
                else:
                    manifest_row["Image Status"] = "invalid_local_path"
                    manifest_row["Error"] = reason

            # 2) manual_images
            if not wrote_image and manual_images.get(i):
                try:
                    jpeg = _normalize_manual_jpeg(manual_images[i])
                    filename = _unique_filename(
                        build_image_filename(brand, sku, product_name),
                        salt=f"{i}:{product_name}",
                    )
                    manifest_row["Local Image Filename"] = filename
                    manifest_row["Image Status"] = "manually_uploaded"
                    manifest_row["Error"] = ""
                    zf.writestr(f"images/{filename}", jpeg)
                    export_image_filename = filename
                    wrote_image = True
                except Exception as exc:
                    manifest_row["Image Status"] = "conversion_error"
                    manifest_row["Error"] = str(exc)

            # 3) remote URL download
            if not wrote_image and _is_public_https_image_url(image_url):
                result = _download_convert(
                    image_url,
                    brand=brand,
                    model_sku=sku,
                    product_name=product_name,
                )
                manifest_row["Image Status"] = result["image_status"]
                manifest_row["Error"] = result.get("error", "")
                if result["image_status"] == "downloaded" and result.get("jpeg_bytes"):
                    filename = _unique_filename(result["local_image_filename"], salt=f"{i}:{product_name}:{image_url}")
                    manifest_row["Local Image Filename"] = filename
                    export_image_filename = filename
                    zf.writestr(f"images/{filename}", result["jpeg_bytes"])

            manifest_rows.append(manifest_row)
            export_image_filenames.append(export_image_filename)

        if export_rows:
            df = pd.DataFrame(export_rows, columns=PROGRAMA_COLUMNS)
            df["Image Filename"] = export_image_filenames
        else:
            df = pd.DataFrame(columns=PROGRAMA_COLUMNS + ["Image Filename"])

        zf.writestr("programa_import.csv", export_programa_csv(df).decode("utf-8"))
        zf.writestr("programa_import.xlsx", export_programa_xlsx(df))

        if manifest_rows:
            manifest_df = pd.DataFrame(manifest_rows, columns=_MANIFEST_COLUMNS)
            mbuf = io.StringIO()
            manifest_df.to_csv(mbuf, index=False)
            zf.writestr("manifest.csv", mbuf.getvalue())

            guide_df = _manual_guide_dataframe(manifest_rows)
            gbuf = io.StringIO()
            guide_df.to_csv(gbuf, index=False)
            zf.writestr("manual_image_upload_guide.csv", gbuf.getvalue())

        zf.writestr("README_manual_image_upload.txt", _manual_image_readme())

    return buf.getvalue()
